from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb2 = load_workbook(filename = 'C:\Python36\Lib\site-packages\openpyxl\workbook\example.xlsx')
Test = wb2['Test']
print(Test['A1'].value)
print(Test['A2'].value)
print(Test['A3'].value)
print(Test['A4'].value)
print(Test['A5'].value)
print(Test['A6'].value)
print(Test['A7'].value)
print(Test['A8'].value)
print(Test['A9'].value)
print(Test['A10'].value)
print(Test['A11'].value)