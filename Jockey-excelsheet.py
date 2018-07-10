from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.jockeyindia.com/")
wait = WebDriverWait(driver, 10)

driver.implicitly_wait(20)
driver.find_element_by_xpath("//span[@class='pull-left']").click()

driver.implicitly_wait(20)

driver.find_element_by_xpath("//a[@id='innermenuid']").click()

driver.implicitly_wait(20)
driver.find_element_by_xpath("//a[@class='menuheading ng-binding ng-scope active']").click()
driver.implicitly_wait(20)
#driver.find_element_by_xpath("//li[@class='tab-li ng-scope active']/a[@class='ng-binding' and 1]").click()
driver.find_element_by_xpath("//a[1]/span[@class='ng-binding ng-scope' and 2]").click()
driver.find_element_by_xpath("//img[@class='ng-scope']").click()
driver.find_element_by_xpath("//div[2]/a[@class='pdp-size-link' and 1]").click()
driver.find_element_by_xpath("//a[@class='pdp-btn-addtocart ladda-button']").click()
#driver.find_element_by_css_selector("div#home a.pdp-btn-buynow.ladda-button > span.ladda-label > span").click()
driver.find_element_by_xpath("//a[@class='pdp-btn-buynow ladda-button']").click()
driver.find_element_by_xpath("//a[@class='btn-cart-pp ladda-button']").click()

wait = WebDriverWait(driver, 10)

wb2 = load_workbook(filename = 'C:\Python36\Lib\site-packages\openpyxl\workbook\example.xlsx')
Test = wb2['Test']
wait = WebDriverWait(driver, 10)

driver.find_element_by_xpath("//input[@class='form-control ng-pristine ng-untouched ng-empty ng-valid-email ng-invalid ng-invalid-required ng-valid-pattern ng-valid-maxlength']").click()
driver.find_element_by_xpath("//input[@class='form-control ng-pristine ng-untouched ng-empty ng-valid-email ng-invalid ng-invalid-required ng-valid-pattern ng-valid-maxlength']").send_keys((Test['A1'].value))

driver.find_element_by_xpath("//input[@class='form-control ng-pristine ng-untouched ng-empty ng-invalid ng-invalid-required ng-valid-pattern ng-valid-maxlength']").click()
driver.find_element_by_xpath("//input[@class='form-control ng-pristine ng-untouched ng-empty ng-invalid ng-invalid-required ng-valid-pattern ng-valid-maxlength']").send_keys((Test['B1'].value))

driver.find_element_by_xpath("//button/span[@class='ladda-label' and 1]").click()




